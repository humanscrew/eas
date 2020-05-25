import openpyxl
import os.path
import os
import xlrd
# import xlwt
import pandas as pd
import re


def datacopy(fwb, fsheet, wb, wbsheet, row_fwb, row_muban):       # 复制凭证模板信息
    fianlsheet = fwb[fsheet]
    tempsheet = wb[wbsheet]
    maxcolumn = tempsheet.max_column
    for j in range(1, maxcolumn + 1):
        fianlsheet.cell(row=row_fwb, column=j).value = tempsheet.cell(row=row_muban, column=j).value


def datamove(fwb, fsheet, frow, fcolumn, wb, wbsheet, wbrow, wbcolumn):       # 写入数据
    fianlsheet = fwb[fsheet]
    tempsheet = wb[wbsheet]
    fianlsheet.cell(row=frow, column=fcolumn).value = tempsheet.cell(row=wbrow, column=wbcolumn).value


def datedeal(shouru_row, shouru_column, mbrow, fcolumn, fcolumn2, name, rowcount):
    if shourusheet.cell(row=shouru_row, column=shouru_column).value:                        # 导入收入数据
        rowcount += 1
        datacopy(finalWB, sheetname, wbmuban, sheetname, rowcount, mbrow)
        datamove(finalWB, sheetname, rowcount, fcolumn, wbshouru, "收入", shouru_row, shouru_column)
        
        fianlSheet.cell(row=rowcount, column=fcolumn2).value = f"{abstract}{name}"     # 修改引入凭证摘要
        fianlSheet.cell(row=rowcount, column=fcolumn2-4).value = f"{month[0]}"         # 修改引入凭证会计期间
        fianlSheet.cell(row=rowcount, column=fcolumn2-5).value = f"{date}"             # 修改引入凭证业务日期
        fianlSheet.cell(row=rowcount, column=fcolumn2-6).value = f"{date}"             # 修改引入凭证记账日期
    return rowcount


def xjdatedeal(xjrowcount,k,srvalue):
    xjfianlSheet.cell(row=xjrowcount, column=2).value = fianlSheet.cell(row=k, column=2).value
    xjfianlSheet.cell(row=xjrowcount, column=3).value = fianlSheet.cell(row=k, column=4).value
    xjfianlSheet.cell(row=xjrowcount, column=5).value = fianlSheet.cell(row=k, column=6).value
    xjfianlSheet.cell(row=xjrowcount, column=11).value = srvalue
    xjfianlSheet.cell(row=xjrowcount, column=12).value = srvalue
    xjfianlSheet.cell(row=xjrowcount, column=13).value = srvalue


finalWB = openpyxl.Workbook()                             # 新建工作簿用来汇总数据
finalWB.create_sheet(index=0, title="凭证")                 # 新建凭证工作表
finalWB.create_sheet(index=1, title="现金流量")             # 新建现金流量工作表

wbmuban = openpyxl.load_workbook("./米林收入凭证模板.xlsx")                              # 载入米林收入凭证模板

rootdir = "./收入日报表"                                    # 提前将收入日报表放在文件夹下
files = os.listdir(rootdir)                                # 获得文件夹下的所有Excel文件
num = len(files)                                           # 获取excel表数量

rowcount = 0                        # 引入凭证行计数器

sheetname = "凭证"
fianlSheet = finalWB[sheetname]
datacopy(finalWB, sheetname, wbmuban, sheetname, 1, 1)                              # 复制凭证模板标题行
rowcount += 1

xjsheetname = "现金流量"
xjfianlSheet = finalWB[xjsheetname]
datacopy(finalWB, xjsheetname, wbmuban, xjsheetname, 1, 1)                              # 复制现金流量模板标题行
xjrowcount = 1

for i in range(0, num):
    wbshouru = openpyxl.load_workbook(f"./收入日报表/{files[i]}", data_only=True)           # 载入收入日报表
    shourusheet = wbshouru["收入"]                                                              # 获取收入日报表
    abstractwords = files[i].split('.')[0]                                                 # 获取摘要信息

    year = re.findall( r'(.*?)年', abstractwords)
    month = re.findall( r'年(.*?)月', abstractwords)
    day = re.findall( r'月(.*?)日', abstractwords)
    date = year[0]+"-"+month[0]+"-"+day[0]          # 获取年月日信息
    
    abstract = f"收{abstractwords}"
    tempcount = rowcount

    rowcount = datedeal(32, 11, 2, 16, 8, "（现金）", rowcount)            # 导入现金收款
    rowcount = datedeal(32, 5, 3, 16, 8, "（POS机）", rowcount)            # 导入POS刷卡收款
    rowcount = datedeal(32, 8, 4, 16, 8, "（微信）", rowcount)             # 导入微信收款
    rowcount = datedeal(18, 25, 5, 16, 8, "（美团挂账）", rowcount)        # 导入美团挂账
    rowcount = datedeal(21, 25, 6, 16, 8, "（携程挂账）", rowcount)        # 导入携程挂账
    rowcount = datedeal(24, 25, 7, 16, 8, "（西旅官网挂账）", rowcount)     # 导入西旅官网挂账
    rowcount = datedeal(27, 25, 8, 16, 8, "（飞猪挂账）", rowcount)         # 导入飞猪挂账
    rowcount = datedeal(30, 25, 9, 16, 8, "（驴妈妈挂账）", rowcount)       # 导入驴妈妈挂账

    rowcount = datedeal(31, 9, 10, 17, 8, "", rowcount)       # 导入门票收入
    rowcount = datedeal(31, 13, 11, 17, 8, "（车票）", rowcount)           # 导入车票收入
    rowcount = datedeal(31, 23, 12, 17, 8, "（商品售卖）", rowcount)        # 导入商品售卖收入
    rowcount = datedeal(31, 21, 13, 17, 8, "（短程游艇）", rowcount)        # 导入短程游艇收入
    rowcount = datedeal(31, 19, 14, 17, 8, "（热气球）", rowcount)        # 导入热气球收入

    rowcount = datedeal(31, 8, 15, 17, 8, "（计提惠民基金）", rowcount)        # 导入惠民基金
    rowcount = datedeal(31, 7, 16, 17, 8, "（计提米林县分成）", rowcount)        # 导入米林县分成

    rowcount = datedeal(32, 15, 17, 16, 8, "（支付POS机手续费）", rowcount)        # 导入POS机手续费
    rowcount = datedeal(32, 15, 18, 17, 8, "（支付POS机手续费）", rowcount)        # 导入POS机手续费

    rowcount = datedeal(32, 19, 19, 16, 8, "（支付微信手续费）", rowcount)        # 导入微信手续费
    rowcount = datedeal(32, 19, 20, 17, 8, "（支付微信手续费）", rowcount)        # 导入微信手续费

    shouru_maxrow = fianlSheet.max_row                  # 修正辅助账摘要、凭证号、原币金额、辅助帐业务日期、到期日
    for k in range(tempcount+1, shouru_maxrow + 1):
        fianlSheet.cell(row=k, column=33).value = fianlSheet.cell(row=k, column=8).value
        fianlSheet.cell(row=k, column=66).value = fianlSheet.cell(row=k, column=3).value
        fianlSheet.cell(row=k, column=67).value = fianlSheet.cell(row=k, column=3).value

        fianlSheet.cell(row=k, column=6).value = f"000{i+1}"
        if fianlSheet.cell(row=k, column=16).value:
            fianlSheet.cell(row=k, column=13).value = fianlSheet.cell(row=k, column=16).value
        else:
            fianlSheet.cell(row=k, column=13).value = fianlSheet.cell(row=k, column=17).value

        if fianlSheet.cell(row=k, column=9).value == "1001.01":                         # 生成现金流量表
            xjrowcount += 1
            datacopy(finalWB, xjsheetname, wbmuban, xjsheetname, xjrowcount, 2)         # 库存现金现金流量
            srvalue = fianlSheet.cell(row=k, column=13).value
            xjdatedeal(xjrowcount,k,srvalue)
        elif fianlSheet.cell(row=k, column=9).value == "1002.01":
            key = re.findall(r'（(.*?)）', fianlSheet.cell(row=k, column=8).value)[0]       #银行存款现金流量
            if key == "POS机":
                xjrowcount += 1
                datacopy(finalWB, xjsheetname, wbmuban, xjsheetname, xjrowcount, 3)
                srvalue = fianlSheet.cell(row=k, column=13).value
                xjdatedeal(xjrowcount,k,srvalue)
            if key == "支付POS机手续费":
                xjrowcount += 1
                datacopy(finalWB, xjsheetname, wbmuban, xjsheetname, xjrowcount, 7)
                srvalue = fianlSheet.cell(row=k, column=13).value
                xjdatedeal(xjrowcount,k,srvalue)
            if key == "支付微信手续费":
                xjrowcount += 1
                datacopy(finalWB, xjsheetname, wbmuban, xjsheetname, xjrowcount, 8)
                srvalue = fianlSheet.cell(row=k, column=13).value
                xjdatedeal(xjrowcount,k,srvalue)
            if key == "微信":
                xjrowcount += 1
                datacopy(finalWB, xjsheetname, wbmuban, xjsheetname, xjrowcount, 4)
                srvalue = shourusheet.cell(row=32, column=8).value-shourusheet.cell(row=31, column=8).value-shourusheet.cell(row=31, column=7).value
                xjdatedeal(xjrowcount,k,srvalue)

                xjrowcount += 1
                datacopy(finalWB, xjsheetname, wbmuban, xjsheetname, xjrowcount, 5)             # 微信部分现金流量
                srvalue = shourusheet.cell(row=31, column=8).value
                xjdatedeal(xjrowcount,k,srvalue)
                
                xjrowcount += 1
                datacopy(finalWB, xjsheetname, wbmuban, xjsheetname, xjrowcount, 6)
                srvalue = shourusheet.cell(row=31, column=7).value
                xjdatedeal(xjrowcount,k,srvalue)

print(rowcount)

finalWB.save(f"./米林收入引入凭证/{abstractwords}引入凭证.xls")  # 保存引入凭证
finalWB.close()

pos1 = pd.read_excel(f"./米林收入引入凭证/{abstractwords}引入凭证.xls", sheet_name='凭证', header=None, index=None)         # 读取凭证工作表信息
pos2 = pd.read_excel(f"./米林收入引入凭证/{abstractwords}引入凭证.xls", sheet_name='现金流量', header=None, index=None)     # 读取现金流量表信息

writer = pd.ExcelWriter(f"./米林收入引入凭证/{abstractwords}引入凭证1.xls")

pos1.to_excel(writer, sheet_name='凭证', header=None, index=None)
pos2.to_excel(writer, sheet_name='现金流量', header=None, index=None)
writer.save()                            # 文件保存
writer.close()                          # 文件关闭
